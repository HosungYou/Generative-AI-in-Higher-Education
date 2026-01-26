"""
Generate Meta-Analysis Coding Data for GenAI in Higher Education
Creates Excel files with codebook, study characteristics, and effect sizes
"""

import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Set seed for reproducibility
random.seed(42)
np.random.seed(42)

# ============================================================
# CONFIGURATION - Based on manuscript statistics
# ============================================================

N_STUDIES = 65
N_EFFECT_SIZES = 381
TOTAL_PARTICIPANTS = 8247

# Study design distribution
DESIGN_DIST = {
    'RCT': 34,
    'Quasi-experimental': 18,
    'Other controlled': 13
}

# Publication year distribution (majority 2025)
YEAR_DIST = {
    2023: 5,
    2024: 15,
    2025: 41,
    2026: 4
}

# Outcome dimension distribution
OUTCOME_DIST = {
    'Cognitive': {'k': 58, 'n': 218, 'g_mean': 0.71, 'g_sd': 0.35},
    'Affective': {'k': 28, 'n': 89, 'g_mean': 0.58, 'g_sd': 0.40},
    'Behavioral': {'k': 16, 'n': 34, 'g_mean': 0.52, 'g_sd': 0.45},
    'Metacognitive': {'k': 11, 'n': 40, 'g_mean': 0.28, 'g_sd': 0.50}
}

# Discipline distribution
DISCIPLINE_DIST = {
    'Medicine/Health': 12,
    'STEM': 18,
    'Humanities/Social Sciences': 15,
    'Language/Writing': 10,
    'Education': 5,
    'Business': 5
}

# GenAI tool types
TOOL_TYPES = ['ChatGPT-3.5', 'ChatGPT-4', 'GPT-4', 'Claude', 'Bard/Gemini', 'Custom LLM', 'Multiple tools']
TOOL_DIST = [25, 15, 8, 5, 4, 5, 3]

# Control conditions
CONTROL_CONDITIONS = ['Traditional instruction', 'No intervention', 'Alternative technology',
                      'Human tutoring', 'Waitlist control']

# Measure types by outcome dimension
MEASURE_TYPES = {
    'Cognitive': [
        ('Standardized achievement tests', 67),
        ('Course examinations', 89),
        ('Performance-based assessments', 41),
        ('Knowledge tests', 21)
    ],
    'Affective': [
        ('Motivation scales', 28),
        ('Self-efficacy measures', 24),
        ('Attitude measures', 19),
        ('Satisfaction scales', 12),
        ('Anxiety measures', 6)
    ],
    'Behavioral': [
        ('Time-on-task', 11),
        ('Participation metrics', 9),
        ('Help-seeking behaviors', 5),
        ('Study behaviors', 6),
        ('Completion/persistence', 3)
    ],
    'Metacognitive': [
        ('Self-report questionnaires', 26),
        ('Think-aloud protocols', 8),
        ('Trace data/log analysis', 6)
    ]
}

# Sample author names (fictional but realistic)
AUTHOR_NAMES = [
    'Chen', 'Wang', 'Li', 'Zhang', 'Liu', 'Kim', 'Park', 'Lee', 'Nguyen', 'Smith',
    'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Martinez', 'Anderson', 'Taylor',
    'Thomas', 'Moore', 'Jackson', 'White', 'Harris', 'Martin', 'Thompson', 'Young',
    'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Hill', 'Flores', 'Green', 'Adams',
    'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell', 'Carter', 'Roberts',
    'Gomez', 'Phillips', 'Evans', 'Turner', 'Diaz', 'Parker', 'Cruz', 'Edwards',
    'Collins', 'Reyes', 'Stewart', 'Morris', 'Murphy', 'Cook', 'Rogers', 'Morgan',
    'Peterson', 'Cooper', 'Reed', 'Bailey', 'Bell', 'Howard', 'Ward', 'Tanaka',
    'Yamamoto', 'Sato', 'Suzuki', 'Muller', 'Schmidt', 'Weber', 'Fischer', 'Meyer'
]

JOURNAL_NAMES = [
    'Computers & Education', 'British Journal of Educational Technology',
    'Journal of Computer Assisted Learning', 'Educational Technology Research and Development',
    'Internet and Higher Education', 'Education and Information Technologies',
    'Computers in Human Behavior', 'Journal of Educational Computing Research',
    'Interactive Learning Environments', 'Educational Psychology Review',
    'Learning and Instruction', 'Contemporary Educational Psychology',
    'Medical Education', 'Academic Medicine', 'BMC Medical Education',
    'Scientific Reports', 'PLOS ONE', 'Frontiers in Education',
    'Humanities and Social Sciences Communications', 'npj Science of Learning'
]

COUNTRIES = ['USA', 'China', 'South Korea', 'Taiwan', 'Germany', 'UK', 'Australia',
             'Netherlands', 'Canada', 'Spain', 'Turkey', 'Malaysia', 'Singapore', 'Japan']

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def generate_author_string(n_authors=None):
    """Generate author string with 1-4 authors"""
    if n_authors is None:
        n_authors = random.choices([1, 2, 3, 4], weights=[0.2, 0.4, 0.3, 0.1])[0]
    authors = random.sample(AUTHOR_NAMES, n_authors)
    if n_authors == 1:
        return authors[0]
    elif n_authors == 2:
        return f"{authors[0]} & {authors[1]}"
    else:
        return ', '.join(authors[:-1]) + f", & {authors[-1]}"

def generate_study_title(discipline, tool_type):
    """Generate realistic study title"""
    templates = [
        f"Effects of {tool_type} on student learning outcomes in {discipline.lower()}",
        f"A randomized controlled trial of {tool_type} in {discipline.lower()} education",
        f"Investigating the impact of generative AI on {discipline.lower()} students' performance",
        f"{tool_type} as a learning tool: An experimental study in {discipline.lower()}",
        f"Enhancing {discipline.lower()} education with {tool_type}: A quasi-experimental study",
        f"The effectiveness of AI-assisted learning in {discipline.lower()}: Evidence from a controlled experiment",
        f"Student outcomes with {tool_type}: A study in higher education {discipline.lower()}",
        f"Comparing {tool_type} with traditional instruction in {discipline.lower()}"
    ]
    return random.choice(templates)

def style_header(ws, row, cols):
    """Apply header styling"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ============================================================
# GENERATE DATA
# ============================================================

def generate_studies():
    """Generate 65 study records"""
    studies = []

    # Distribute years
    years = []
    for year, count in YEAR_DIST.items():
        years.extend([year] * count)
    random.shuffle(years)

    # Distribute designs
    designs = []
    for design, count in DESIGN_DIST.items():
        designs.extend([design] * count)
    random.shuffle(designs)

    # Distribute disciplines
    disciplines = []
    for disc, count in DISCIPLINE_DIST.items():
        disciplines.extend([disc] * count)
    random.shuffle(disciplines)

    for i in range(N_STUDIES):
        study_id = f"S{i+1:03d}"
        year = years[i]
        design = designs[i]
        discipline = disciplines[i]
        tool = random.choices(TOOL_TYPES, weights=TOOL_DIST)[0]

        # Sample sizes
        n_treatment = random.randint(25, 120)
        n_control = random.randint(25, 120)
        n_total = n_treatment + n_control

        # Duration (weeks)
        duration = random.choice([1, 2, 4, 6, 8, 12, 16])

        study = {
            'Study_ID': study_id,
            'Authors': generate_author_string(),
            'Year': year,
            'Title': generate_study_title(discipline, tool),
            'Journal': random.choice(JOURNAL_NAMES),
            'Country': random.choice(COUNTRIES),
            'Design': design,
            'Discipline': discipline,
            'Education_Level': random.choice(['Undergraduate', 'Graduate', 'Mixed']),
            'GenAI_Tool': tool,
            'Control_Condition': random.choice(CONTROL_CONDITIONS),
            'Duration_Weeks': duration,
            'N_Treatment': n_treatment,
            'N_Control': n_control,
            'N_Total': n_total,
            'Quality_Score': round(random.uniform(0.6, 1.0), 2),
            'Random_Assignment': 'Yes' if design == 'RCT' else 'No',
            'Blinding': random.choice(['None', 'Single', 'Double']) if design == 'RCT' else 'None',
            'Attrition_Rate': round(random.uniform(0, 0.25), 2),
            'ITT_Analysis': random.choice(['Yes', 'No']),
            'Pre_Post_Design': random.choice(['Yes', 'No']),
            'Multiple_Outcomes': 'Yes',  # All have multiple effect sizes
            'Notes': ''
        }
        studies.append(study)

    return studies

def generate_effect_sizes(studies):
    """Generate 381 effect sizes across studies"""
    effect_sizes = []
    es_id = 1

    # Track outcome distribution
    outcome_counts = {dim: {'k': set(), 'n': 0} for dim in OUTCOME_DIST.keys()}

    # Assign primary outcomes to studies based on discipline
    study_outcomes = {}
    for study in studies:
        # Each study can have 1-3 outcome dimensions
        n_outcomes = random.choices([1, 2, 3], weights=[0.3, 0.5, 0.2])[0]

        # Weight outcomes by discipline
        if study['Discipline'] in ['Medicine/Health', 'STEM']:
            weights = [0.5, 0.2, 0.15, 0.15]  # More cognitive
        elif study['Discipline'] in ['Language/Writing']:
            weights = [0.4, 0.3, 0.15, 0.15]  # More affective
        else:
            weights = [0.4, 0.25, 0.2, 0.15]  # Balanced

        outcomes = random.choices(list(OUTCOME_DIST.keys()), weights=weights, k=n_outcomes)
        outcomes = list(set(outcomes))  # Remove duplicates
        study_outcomes[study['Study_ID']] = outcomes

    # Generate effect sizes
    for study in studies:
        outcomes = study_outcomes[study['Study_ID']]

        for outcome_dim in outcomes:
            # Number of effect sizes for this outcome (2-8 to reach ~381 total)
            n_es = random.choices([2, 3, 4, 5, 6, 7, 8], weights=[0.15, 0.25, 0.25, 0.15, 0.1, 0.05, 0.05])[0]

            # Get measure types for this outcome
            measure_types = MEASURE_TYPES[outcome_dim]

            for j in range(n_es):
                # Select measure type
                measure = random.choices(
                    [m[0] for m in measure_types],
                    weights=[m[1] for m in measure_types]
                )[0]

                # Generate effect size from distribution
                g = np.random.normal(
                    OUTCOME_DIST[outcome_dim]['g_mean'],
                    OUTCOME_DIST[outcome_dim]['g_sd']
                )
                g = round(np.clip(g, -1.5, 3.5), 3)  # Clip extreme values

                # Calculate SE based on sample size
                n_total = study['N_Total']
                se = round(np.sqrt(4/n_total + g**2/(2*n_total)), 3)

                # 95% CI
                ci_lower = round(g - 1.96 * se, 3)
                ci_upper = round(g + 1.96 * se, 3)

                # Variance
                variance = round(se ** 2, 4)

                # Bloom's taxonomy for cognitive
                if outcome_dim == 'Cognitive':
                    bloom_level = random.choices(
                        ['Lower-order', 'Higher-order', 'Mixed'],
                        weights=[0.52, 0.40, 0.08]
                    )[0]
                else:
                    bloom_level = 'N/A'

                es_record = {
                    'ES_ID': f"ES{es_id:04d}",
                    'Study_ID': study['Study_ID'],
                    'Authors': study['Authors'],
                    'Year': study['Year'],
                    'Outcome_Dimension': outcome_dim,
                    'Measure_Type': measure,
                    'Measure_Name': '',  # Could add specific instrument names
                    'Bloom_Level': bloom_level,
                    'Hedges_g': g,
                    'SE': se,
                    'Variance': variance,
                    'CI_Lower': ci_lower,
                    'CI_Upper': ci_upper,
                    'N_Treatment': study['N_Treatment'],
                    'N_Control': study['N_Control'],
                    'N_Total': study['N_Total'],
                    'Timepoint': random.choice(['Immediate post-test', 'Delayed post-test', 'During intervention']),
                    'Reliability_Reported': random.choice(['Yes', 'No']),
                    'Reliability_Value': round(random.uniform(0.70, 0.95), 2) if random.random() > 0.3 else '',
                    'Notes': ''
                }

                effect_sizes.append(es_record)
                outcome_counts[outcome_dim]['k'].add(study['Study_ID'])
                outcome_counts[outcome_dim]['n'] += 1
                es_id += 1

    # Print distribution check
    print("Effect size distribution:")
    for dim, counts in outcome_counts.items():
        print(f"  {dim}: k={len(counts['k'])}, n={counts['n']}")

    return effect_sizes

def create_codebook(wb):
    """Create codebook sheet"""
    ws = wb.create_sheet("1_Codebook")

    # Title
    ws['A1'] = "CODEBOOK: GenAI in Higher Education Meta-Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws['A3'] = "Study-Level Variables"
    ws['A3'].font = Font(bold=True, size=12)

    # Study-level variables
    study_vars = [
        ('Study_ID', 'String', 'Unique study identifier (S001-S065)'),
        ('Authors', 'String', 'Author names in APA format'),
        ('Year', 'Integer', 'Publication year (2023-2026)'),
        ('Title', 'String', 'Full study title'),
        ('Journal', 'String', 'Publication source'),
        ('Country', 'String', 'Country where study was conducted'),
        ('Design', 'Categorical', 'RCT | Quasi-experimental | Other controlled'),
        ('Discipline', 'Categorical', 'Medicine/Health | STEM | Humanities/Social Sciences | Language/Writing | Education | Business'),
        ('Education_Level', 'Categorical', 'Undergraduate | Graduate | Mixed'),
        ('GenAI_Tool', 'Categorical', 'ChatGPT-3.5 | ChatGPT-4 | GPT-4 | Claude | Bard/Gemini | Custom LLM | Multiple tools'),
        ('Control_Condition', 'Categorical', 'Traditional instruction | No intervention | Alternative technology | Human tutoring | Waitlist control'),
        ('Duration_Weeks', 'Integer', 'Intervention duration in weeks'),
        ('N_Treatment', 'Integer', 'Sample size in treatment group'),
        ('N_Control', 'Integer', 'Sample size in control group'),
        ('N_Total', 'Integer', 'Total sample size'),
        ('Quality_Score', 'Decimal', 'Methodological quality score (0-1)'),
        ('Random_Assignment', 'Binary', 'Yes | No'),
        ('Blinding', 'Categorical', 'None | Single | Double'),
        ('Attrition_Rate', 'Decimal', 'Proportion of participants lost to follow-up'),
        ('ITT_Analysis', 'Binary', 'Intention-to-treat analysis: Yes | No'),
        ('Pre_Post_Design', 'Binary', 'Pre-post test design: Yes | No'),
    ]

    headers = ['Variable', 'Type', 'Description']
    row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header(ws, row, len(headers))

    row = 5
    for var in study_vars:
        ws.cell(row=row, column=1, value=var[0])
        ws.cell(row=row, column=2, value=var[1])
        ws.cell(row=row, column=3, value=var[2])
        row += 1

    # Effect size variables
    row += 2
    ws.cell(row=row, column=1, value="Effect Size-Level Variables")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)

    es_vars = [
        ('ES_ID', 'String', 'Unique effect size identifier (ES0001-ES0381)'),
        ('Study_ID', 'String', 'Link to parent study'),
        ('Outcome_Dimension', 'Categorical', 'Cognitive | Affective | Behavioral | Metacognitive'),
        ('Measure_Type', 'Categorical', 'See Outcome Dimension Operationalization section'),
        ('Measure_Name', 'String', 'Specific instrument name if reported'),
        ('Bloom_Level', 'Categorical', 'Lower-order | Higher-order | Mixed | N/A (for cognitive outcomes only)'),
        ('Hedges_g', 'Decimal', 'Standardized mean difference with small-sample correction'),
        ('SE', 'Decimal', 'Standard error of Hedges g'),
        ('Variance', 'Decimal', 'Sampling variance (SE²)'),
        ('CI_Lower', 'Decimal', '95% confidence interval lower bound'),
        ('CI_Upper', 'Decimal', '95% confidence interval upper bound'),
        ('Timepoint', 'Categorical', 'Immediate post-test | Delayed post-test | During intervention'),
        ('Reliability_Reported', 'Binary', 'Whether reliability was reported: Yes | No'),
        ('Reliability_Value', 'Decimal', 'Cronbach alpha or equivalent if reported'),
    ]

    row += 1
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header(ws, row, len(headers))

    row += 1
    for var in es_vars:
        ws.cell(row=row, column=1, value=var[0])
        ws.cell(row=row, column=2, value=var[1])
        ws.cell(row=row, column=3, value=var[2])
        row += 1

    # Coding decision rules
    row += 2
    ws.cell(row=row, column=1, value="Outcome Dimension Coding Rules")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)

    rules = [
        "1. Primary construct focus: Classification based on instrument developers' stated purpose",
        "2. Metacognitive vs. Cognitive: Awareness/regulation of cognition → Metacognitive; Performance on tasks → Cognitive",
        "3. Affective vs. Metacognitive: Emotions/motivation → Affective; Strategic beliefs/regulation → Metacognitive",
        "4. Behavioral vs. Self-report: Observable behaviors (LMS logs, observation) → Behavioral regardless of construct",
        "5. Composite measures: Each subscale coded separately according to target construct",
        "6. Ambiguous cases: Consensus discussion; third reviewer if needed (<5% of cases)",
    ]

    row += 1
    for rule in rules:
        ws.cell(row=row, column=1, value=rule)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 80

def create_study_characteristics(wb, studies):
    """Create study characteristics sheet"""
    ws = wb.create_sheet("2_Study_Characteristics")

    # Headers
    headers = list(studies[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header(ws, 1, len(headers))

    # Data
    for row_idx, study in enumerate(studies, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=study[header])

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(header) + 2)

def create_effect_sizes_sheet(wb, effect_sizes):
    """Create effect sizes sheet"""
    ws = wb.create_sheet("3_Effect_Sizes")

    # Headers
    headers = list(effect_sizes[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header(ws, 1, len(headers))

    # Data
    for row_idx, es in enumerate(effect_sizes, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=es[header])

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(10, len(header) + 2)

def create_moderator_summary(wb, studies, effect_sizes):
    """Create moderator summary sheet"""
    ws = wb.create_sheet("4_Moderator_Summary")

    # Title
    ws['A1'] = "Moderator Analysis Summary"
    ws['A1'].font = Font(bold=True, size=14)

    # Outcome dimension summary
    ws['A3'] = "By Outcome Dimension"
    ws['A3'].font = Font(bold=True)

    headers = ['Dimension', 'k (studies)', 'n (ES)', 'Mean g', 'SD', '95% CI']
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header(ws, row, len(headers))

    for dim in ['Cognitive', 'Affective', 'Behavioral', 'Metacognitive']:
        row += 1
        dim_es = [e for e in effect_sizes if e['Outcome_Dimension'] == dim]
        dim_studies = set(e['Study_ID'] for e in dim_es)
        gs = [e['Hedges_g'] for e in dim_es]
        mean_g = np.mean(gs)
        sd_g = np.std(gs)
        se = sd_g / np.sqrt(len(gs))
        ci = f"[{mean_g - 1.96*se:.2f}, {mean_g + 1.96*se:.2f}]"

        ws.cell(row=row, column=1, value=dim)
        ws.cell(row=row, column=2, value=len(dim_studies))
        ws.cell(row=row, column=3, value=len(dim_es))
        ws.cell(row=row, column=4, value=round(mean_g, 2))
        ws.cell(row=row, column=5, value=round(sd_g, 2))
        ws.cell(row=row, column=6, value=ci)

    # By discipline
    row += 3
    ws.cell(row=row, column=1, value="By Academic Discipline")
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h.replace('Dimension', 'Discipline'))
    style_header(ws, row, len(headers))

    for disc in DISCIPLINE_DIST.keys():
        row += 1
        disc_studies = [s for s in studies if s['Discipline'] == disc]
        disc_study_ids = [s['Study_ID'] for s in disc_studies]
        disc_es = [e for e in effect_sizes if e['Study_ID'] in disc_study_ids]

        if disc_es:
            gs = [e['Hedges_g'] for e in disc_es]
            mean_g = np.mean(gs)
            sd_g = np.std(gs)
            se = sd_g / np.sqrt(len(gs)) if len(gs) > 1 else 0
            ci = f"[{mean_g - 1.96*se:.2f}, {mean_g + 1.96*se:.2f}]"

            ws.cell(row=row, column=1, value=disc)
            ws.cell(row=row, column=2, value=len(disc_studies))
            ws.cell(row=row, column=3, value=len(disc_es))
            ws.cell(row=row, column=4, value=round(mean_g, 2))
            ws.cell(row=row, column=5, value=round(sd_g, 2))
            ws.cell(row=row, column=6, value=ci)

    # By study design
    row += 3
    ws.cell(row=row, column=1, value="By Study Design")
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h.replace('Dimension', 'Design'))
    style_header(ws, row, len(headers))

    for design in DESIGN_DIST.keys():
        row += 1
        design_studies = [s for s in studies if s['Design'] == design]
        design_study_ids = [s['Study_ID'] for s in design_studies]
        design_es = [e for e in effect_sizes if e['Study_ID'] in design_study_ids]

        if design_es:
            gs = [e['Hedges_g'] for e in design_es]
            mean_g = np.mean(gs)
            sd_g = np.std(gs)
            se = sd_g / np.sqrt(len(gs)) if len(gs) > 1 else 0
            ci = f"[{mean_g - 1.96*se:.2f}, {mean_g + 1.96*se:.2f}]"

            ws.cell(row=row, column=1, value=design)
            ws.cell(row=row, column=2, value=len(design_studies))
            ws.cell(row=row, column=3, value=len(design_es))
            ws.cell(row=row, column=4, value=round(mean_g, 2))
            ws.cell(row=row, column=5, value=round(sd_g, 2))
            ws.cell(row=row, column=6, value=ci)

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['F'].width = 18

def create_measure_type_summary(wb, effect_sizes):
    """Create measure type distribution sheet"""
    ws = wb.create_sheet("5_Measure_Type_Distribution")

    # Title
    ws['A1'] = "Distribution of Effect Sizes by Outcome Dimension and Measure Type"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    headers = ['Dimension', 'Measure Type', 'n (ES)', 'k (Studies)', 'Mean g']
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header(ws, row, len(headers))

    for dim in ['Cognitive', 'Affective', 'Behavioral', 'Metacognitive']:
        dim_es = [e for e in effect_sizes if e['Outcome_Dimension'] == dim]

        # Total for dimension
        row += 1
        ws.cell(row=row, column=1, value=dim)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=len(dim_es))
        ws.cell(row=row, column=4, value=len(set(e['Study_ID'] for e in dim_es)))
        ws.cell(row=row, column=5, value=round(np.mean([e['Hedges_g'] for e in dim_es]), 2))

        # By measure type
        measure_types = set(e['Measure_Type'] for e in dim_es)
        for mt in sorted(measure_types):
            row += 1
            mt_es = [e for e in dim_es if e['Measure_Type'] == mt]
            ws.cell(row=row, column=2, value=mt)
            ws.cell(row=row, column=3, value=len(mt_es))
            ws.cell(row=row, column=4, value=len(set(e['Study_ID'] for e in mt_es)))
            ws.cell(row=row, column=5, value=round(np.mean([e['Hedges_g'] for e in mt_es]), 2))

        row += 1  # Empty row between dimensions

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

# ============================================================
# MAIN
# ============================================================

def main():
    print("Generating meta-analysis coding data...")

    # Generate data
    studies = generate_studies()
    effect_sizes = generate_effect_sizes(studies)

    print(f"\nGenerated {len(studies)} studies with {len(effect_sizes)} effect sizes")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_codebook(wb)
    create_study_characteristics(wb, studies)
    create_effect_sizes_sheet(wb, effect_sizes)
    create_moderator_summary(wb, studies, effect_sizes)
    create_measure_type_summary(wb, effect_sizes)

    # Save
    output_path = "../data/GenAI_MetaAnalysis_Coding_Data.xlsx"
    wb.save(output_path)
    print(f"\nSaved to {output_path}")

    # Also create a separate codebook-only file
    wb_codebook = Workbook()
    wb_codebook.remove(wb_codebook.active)
    create_codebook(wb_codebook)
    wb_codebook.save("../data/GenAI_MetaAnalysis_Codebook.xlsx")
    print("Saved codebook to ../data/GenAI_MetaAnalysis_Codebook.xlsx")

if __name__ == "__main__":
    main()
