#!/usr/bin/env python3
"""
Convert Actual Meta-Analysis Extraction Data to Excel

This script reads the REAL extracted data from the ScholaRAG meta-analysis
project and converts it to properly formatted Excel files.

Data Source: ScholaRAG/projects/2025-12-05_GenAI-Learning-Effects-Meta/data/07_meta_analysis/
- meta_analysis_effects_final_complete.csv (372 effect sizes)
- Analysis from Codex/meta_analysis_effects_unified_with_moderators_noderived.csv (with moderators)

Author: Hosung You
Date: 2026-01-22
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Paths
SOURCE_DIR = '/Volumes/External SSD/Projects/Research/Done/January/GenAI_Effectiveness/ScholaRAG/projects/2025-12-05_GenAI-Learning-Effects-Meta/data/07_meta_analysis'
OUTPUT_DIR = '/Volumes/External SSD/Projects/Research/Done/January/GenAI_Effectiveness/Generative-AI-in-Higher-Education/data'

# Style settings
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_data():
    """Load the actual extraction data."""
    print("[1] Loading actual extraction data...")

    # Try to load the file with moderators first (more complete)
    moderator_file = os.path.join(SOURCE_DIR, 'Analysis from Codex',
                                   'meta_analysis_effects_unified_with_moderators_noderived.csv')
    basic_file = os.path.join(SOURCE_DIR, 'meta_analysis_effects_final_complete.csv')

    if os.path.exists(moderator_file):
        df = pd.read_csv(moderator_file)
        print(f"    Loaded: {moderator_file}")
        print(f"    Effect sizes: {len(df)}")
    elif os.path.exists(basic_file):
        df = pd.read_csv(basic_file)
        print(f"    Loaded: {basic_file}")
        print(f"    Effect sizes: {len(df)}")
    else:
        raise FileNotFoundError("Could not find extraction data files")

    return df

def create_study_characteristics(df):
    """Extract unique study-level characteristics."""
    print("[2] Extracting study-level characteristics...")

    # Get unique studies
    studies = df.groupby('study_id').first().reset_index()

    # Select and rename columns for study characteristics
    study_cols = {
        'study_id': 'Study_ID',
        'title': 'Title',
        'year': 'Year',
        'authors': 'Authors',
        'study_design': 'Study_Design',
        'genai_tool': 'GenAI_Tool',
        'source': 'Database_Source',
        'doi': 'DOI'
    }

    # Add sample size (average across outcomes for each study)
    sample_sizes = df.groupby('study_id').agg({
        'n_treatment': 'first',
        'n_control': 'first'
    }).reset_index()
    sample_sizes['Sample_Size'] = sample_sizes['n_treatment'] + sample_sizes['n_control']

    studies = studies.merge(sample_sizes[['study_id', 'Sample_Size']], on='study_id', how='left')

    # Calculate number of effect sizes per study
    es_counts = df.groupby('study_id').size().reset_index(name='n_Effects')
    studies = studies.merge(es_counts, on='study_id', how='left')

    # Select columns that exist
    available_cols = [c for c in study_cols.keys() if c in studies.columns]
    result = studies[available_cols + ['Sample_Size', 'n_Effects']].copy()
    result = result.rename(columns=study_cols)

    print(f"    Studies extracted: {len(result)}")
    return result

def create_effect_sizes_sheet(df):
    """Prepare effect sizes data for Excel."""
    print("[3] Preparing effect sizes data...")

    # Select key columns for effect sizes sheet
    es_cols = [
        'study_id', 'outcome_id', 'title', 'year', 'authors',
        'outcome_name', 'outcome_dimension', 'blooms_level',
        'n_treatment', 'n_control',
        'm_treatment', 'sd_treatment', 'm_control', 'sd_control',
        'hedges_g', 'se_g', 'ci_lower', 'ci_upper',
        'cohens_d', 't_value', 'f_value', 'p_value',
        'effect_source', 'extraction_method', 'extraction_confidence',
        'study_design', 'genai_tool', 'measure_type', 'measure_instrument'
    ]

    # Only keep columns that exist
    available_cols = [c for c in es_cols if c in df.columns]
    result = df[available_cols].copy()

    # Rename columns for clarity
    rename_map = {
        'study_id': 'Study_ID',
        'outcome_id': 'ES_ID',
        'title': 'Title',
        'year': 'Year',
        'authors': 'Authors',
        'outcome_name': 'Outcome_Name',
        'outcome_dimension': 'Outcome_Dimension',
        'blooms_level': 'Blooms_Level',
        'n_treatment': 'n_Treatment',
        'n_control': 'n_Control',
        'm_treatment': 'M_Treatment',
        'sd_treatment': 'SD_Treatment',
        'm_control': 'M_Control',
        'sd_control': 'SD_Control',
        'hedges_g': 'Hedges_g',
        'se_g': 'SE_g',
        'ci_lower': 'CI_Lower',
        'ci_upper': 'CI_Upper',
        'cohens_d': 'Cohens_d',
        't_value': 't_value',
        'f_value': 'F_value',
        'p_value': 'p_value',
        'effect_source': 'Effect_Source',
        'extraction_method': 'Extraction_Method',
        'extraction_confidence': 'Extraction_Confidence',
        'study_design': 'Study_Design',
        'genai_tool': 'GenAI_Tool',
        'measure_type': 'Measure_Type',
        'measure_instrument': 'Measure_Instrument'
    }
    result = result.rename(columns={k: v for k, v in rename_map.items() if k in result.columns})

    print(f"    Effect sizes: {len(result)}")
    return result

def create_moderator_summary(df):
    """Calculate moderator summary statistics."""
    print("[4] Calculating moderator summary...")

    summaries = []

    # Outcome Dimension
    for dim in df['outcome_dimension'].dropna().unique():
        subset = df[df['outcome_dimension'] == dim]
        summaries.append({
            'Moderator': 'Outcome Dimension',
            'Category': dim,
            'k_Studies': subset['study_id'].nunique(),
            'n_Effects': len(subset),
            'Mean_g': subset['hedges_g'].mean(),
            'SD_g': subset['hedges_g'].std(),
            'Min_g': subset['hedges_g'].min(),
            'Max_g': subset['hedges_g'].max()
        })

    # GenAI Tool
    if 'genai_tool' in df.columns:
        for tool in df['genai_tool'].dropna().unique():
            if pd.notna(tool) and tool != 'not_reported':
                subset = df[df['genai_tool'] == tool]
                summaries.append({
                    'Moderator': 'GenAI Tool',
                    'Category': tool,
                    'k_Studies': subset['study_id'].nunique(),
                    'n_Effects': len(subset),
                    'Mean_g': subset['hedges_g'].mean(),
                    'SD_g': subset['hedges_g'].std(),
                    'Min_g': subset['hedges_g'].min(),
                    'Max_g': subset['hedges_g'].max()
                })

    # Study Design
    if 'study_design' in df.columns:
        for design in df['study_design'].dropna().unique():
            if pd.notna(design) and design != 'not_reported':
                subset = df[df['study_design'] == design]
                summaries.append({
                    'Moderator': 'Study Design',
                    'Category': design,
                    'k_Studies': subset['study_id'].nunique(),
                    'n_Effects': len(subset),
                    'Mean_g': subset['hedges_g'].mean(),
                    'SD_g': subset['hedges_g'].std(),
                    'Min_g': subset['hedges_g'].min(),
                    'Max_g': subset['hedges_g'].max()
                })

    result = pd.DataFrame(summaries)

    # Round numeric columns
    for col in ['Mean_g', 'SD_g', 'Min_g', 'Max_g']:
        if col in result.columns:
            result[col] = result[col].round(4)

    print(f"    Moderator categories: {len(result)}")
    return result

def create_codebook():
    """Create codebook with variable definitions."""
    print("[5] Creating codebook...")

    codebook_data = [
        ['Variable', 'Description', 'Values/Coding Rules', 'Source'],
        ['Study_ID', 'Unique study identifier', 'Integer (1-69)', 'Assigned during screening'],
        ['ES_ID', 'Effect size identifier', 'Format: StudyID_##', 'Assigned during extraction'],
        ['Title', 'Full study title', 'Text', 'Primary study'],
        ['Year', 'Publication year', '2018-2025', 'Primary study'],
        ['Authors', 'All authors (semicolon separated)', 'Text', 'Primary study'],
        ['Outcome_Name', 'Specific outcome measured', 'Text (e.g., Post-test, Course Satisfaction)', 'Primary study'],
        ['Outcome_Dimension', 'Outcome category', 'cognitive, affective, behavioral, metacognitive', 'Coded'],
        ['Blooms_Level', 'Bloom\'s taxonomy level', 'lower_order, higher_order, mixed, not_reported', 'Coded'],
        ['n_Treatment', 'Treatment group sample size', 'Integer', 'Extracted from tables/text'],
        ['n_Control', 'Control group sample size', 'Integer', 'Extracted from tables/text'],
        ['M_Treatment', 'Treatment group mean', 'Numeric', 'Extracted from tables/text'],
        ['SD_Treatment', 'Treatment group SD', 'Numeric', 'Extracted from tables/text'],
        ['M_Control', 'Control group mean', 'Numeric', 'Extracted from tables/text'],
        ['SD_Control', 'Control group SD', 'Numeric', 'Extracted from tables/text'],
        ['Hedges_g', 'Bias-corrected standardized effect size', 'Numeric (calculated)', 'Calculated'],
        ['SE_g', 'Standard error of Hedges\' g', 'Numeric (calculated)', 'Calculated'],
        ['CI_Lower', '95% CI lower bound', 'Numeric', 'Calculated'],
        ['CI_Upper', '95% CI upper bound', 'Numeric', 'Calculated'],
        ['Cohens_d', 'Cohen\'s d (uncorrected)', 'Numeric', 'Calculated'],
        ['t_value', 't-statistic (if reported)', 'Numeric or NA', 'Extracted'],
        ['F_value', 'F-statistic (if reported)', 'Numeric or NA', 'Extracted'],
        ['p_value', 'p-value', 'Numeric (0-1)', 'Extracted'],
        ['Effect_Source', 'How effect was derived', 'calculated_from_table, reported_directly, converted_from_t', 'Coded'],
        ['Extraction_Method', 'Data extraction method', 'table_ocr, manual, rag_extraction', 'Coded'],
        ['Extraction_Confidence', 'Confidence in extraction', 'high, medium, low, pending_review', 'Coded'],
        ['Study_Design', 'Research design', 'RCT, quasi-experimental, not_reported', 'Coded'],
        ['GenAI_Tool', 'AI tool used in study', 'ChatGPT, GPT-3.5, GPT-4, Gemini, Custom LLM, etc.', 'Extracted'],
        ['Measure_Type', 'Type of measurement', 'standardized_test, course_exam, survey, etc.', 'Coded'],
        ['Measure_Instrument', 'Specific instrument name', 'Text (e.g., MSLQ, Custom quiz)', 'Extracted'],
        ['DOI', 'Digital Object Identifier', 'DOI string or NA', 'Extracted'],
        ['Database_Source', 'Source database', 'Semantic Scholar, OpenAlex, arXiv, ERIC, etc.', 'Recorded during search']
    ]

    return codebook_data

def write_to_excel(df, studies_df, es_df, moderator_df, codebook_data):
    """Write all data to Excel files."""
    print("[6] Writing to Excel files...")

    # Main coding data file
    wb = Workbook()

    # Sheet 1: Codebook
    ws = wb.active
    ws.title = '1_Codebook'

    for row_idx, row_data in enumerate(codebook_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25

    # Sheet 2: Study Characteristics
    ws = wb.create_sheet('2_Study_Characteristics')

    for col_idx, col_name in enumerate(studies_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(studies_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

    for col in range(1, len(studies_df.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['B'].width = 80  # Title column
    ws.column_dimensions['D'].width = 50  # Authors column

    # Sheet 3: Effect Sizes
    ws = wb.create_sheet('3_Effect_Sizes')

    for col_idx, col_name in enumerate(es_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(es_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            # Handle NaN values
            if pd.isna(value):
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

    for col in range(1, len(es_df.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 4: Moderator Summary
    ws = wb.create_sheet('4_Moderator_Summary')

    for col_idx, col_name in enumerate(moderator_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(moderator_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

    for col in range(1, len(moderator_df.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save
    output_path = os.path.join(OUTPUT_DIR, 'GenAI_MetaAnalysis_Coding_Data.xlsx')
    wb.save(output_path)
    print(f"    Saved: {output_path}")

    # Also save raw CSV for reproducibility
    csv_path = os.path.join(OUTPUT_DIR, 'GenAI_MetaAnalysis_Effects_Raw.csv')
    df.to_csv(csv_path, index=False)
    print(f"    Saved: {csv_path}")

    return output_path

def main():
    print("="*70)
    print("Converting REAL Meta-Analysis Data to Excel")
    print("="*70)

    # Load actual data
    df = load_data()

    # Create datasets
    studies_df = create_study_characteristics(df)
    es_df = create_effect_sizes_sheet(df)
    moderator_df = create_moderator_summary(df)
    codebook_data = create_codebook()

    # Write to Excel
    output_path = write_to_excel(df, studies_df, es_df, moderator_df, codebook_data)

    print("\n" + "="*70)
    print("Conversion Complete!")
    print(f"Studies: {len(studies_df)}")
    print(f"Effect sizes: {len(es_df)}")
    print(f"Data source: REAL extraction from ScholaRAG")
    print("="*70)

if __name__ == '__main__':
    main()
