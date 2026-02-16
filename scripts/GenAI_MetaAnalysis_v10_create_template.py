#!/usr/bin/env python3
"""
GenAI Meta-Analysis v10 Template Generator
Generates GenAI_MetaAnalysis_v10_TEMPLATE.xlsx with 9 professionally formatted sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

OUTPUT_PATH = "/Users/hosung/GenAI_MetaAnalysis_v10_TEMPLATE.xlsx"

# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TAB_COLORS = {
    "CODEBOOK": "1F4E79",
    "STUDY_CHAR": "00B050",
    "EFFECT_SIZES": "FF0000",
    "RAW_EXTRACTION": "ED7D31",
    "ES_CALCULATIONS": "7030A0",
    "AI_CODING": "00B0F0",
    "DISCREPANCY_LOG": "FFD966",
    "EXCLUSION_LOG": "A6A6A6",
    "SEARCH_LOG": "5B9BD5",
}

# ---------------------------------------------------------------------------
# Sheet definitions: list of (col_name, type, valid_values_or_None, coding_rule, example, notes)
# ---------------------------------------------------------------------------

STUDY_CHAR_VARS = [
    # Identification
    ("study_id", "integer", None, "Unique numeric ID per study", "1", "Sequential integer starting from 1"),
    ("first_author", "string", None, "Last name of the first author", "Smith", ""),
    ("year", "integer", None, "Publication year", "2024", ""),
    ("title", "string", None, "Full title of the paper", "Effects of ChatGPT on Learning", ""),
    ("doi", "string", None, "Digital Object Identifier", "10.1234/abcd.2024.001", "If available"),
    ("source_type", "categorical", "journal|conference|dissertation|preprint", "Type of publication venue", "journal", ""),
    ("search_source", "string", None, "Database or source where the study was found", "Scopus", "e.g. Scopus, WoS, PubMed, manual"),
    # Study Design
    ("study_design", "categorical", "RCT|quasi_experimental", "Randomized or quasi-experimental", "RCT", ""),
    ("random_assignment", "categorical", "yes|no|unclear", "Whether participants were randomly assigned", "yes", ""),
    ("matching_method", "categorical", "none|propensity|stratified|other", "Method used to match groups if not RCT", "none", ""),
    ("blinding", "categorical", "none|single|double|unclear", "Level of blinding", "none", "Blinding is rare in educational studies"),
    ("attrition_reported", "categorical", "yes|no", "Whether attrition/dropout was reported", "yes", ""),
    # Sample
    ("n_total", "integer", None, "Total sample size across all groups", "120", ""),
    ("academic_level", "categorical", "undergraduate|graduate|mixed", "Academic level of participants", "undergraduate", ""),
    ("discipline", "categorical", "CS|education|language|medicine|STEM_other|social_science|other", "Subject discipline", "CS", ""),
    ("country", "string", None, "Country where study was conducted", "USA", "Use ISO-style short names"),
    ("institution_type", "categorical", "public|private|unclear", "Type of institution", "public", ""),
    ("prior_knowledge_controlled", "categorical", "yes|no", "Whether prior knowledge was controlled for", "yes", "Via pretest or covariate"),
    # Intervention
    ("genai_tool", "categorical", "ChatGPT|GPT-3.5|GPT-4|GPT-4o|Claude|Gemini|Copilot|Custom_LLM|Other|Not_specified", "Generative AI tool used", "ChatGPT", ""),
    ("genai_version_detail", "string", None, "Specific version or model details", "GPT-4-turbo-2024-01", "As reported by authors"),
    ("genai_role", "categorical", "tutor|feedback_provider|writing_assistant|problem_solver|content_generator|other", "Role of GenAI in the intervention", "tutor", ""),
    ("control_condition", "categorical", "no_AI|traditional_instruction|alternative_AI|waitlist|other", "What the control group received", "traditional_instruction", ""),
    ("duration_weeks", "numeric", None, "Duration of intervention in weeks", "4", "Use decimals if needed"),
    ("total_sessions", "integer", None, "Total number of intervention sessions", "8", ""),
    ("implementation_context", "categorical", "in_class|homework|both|lab|online", "Where the intervention took place", "in_class", ""),
    # RoB 2.0
    ("rob_randomization", "categorical", "low|some_concerns|high", "RoB domain 1: randomization process", "low", ""),
    ("rob_deviations", "categorical", "low|some_concerns|high", "RoB domain 2: deviations from intended interventions", "low", ""),
    ("rob_missing_data", "categorical", "low|some_concerns|high", "RoB domain 3: missing outcome data", "low", ""),
    ("rob_measurement", "categorical", "low|some_concerns|high", "RoB domain 4: measurement of the outcome", "some_concerns", ""),
    ("rob_selection", "categorical", "low|some_concerns|high", "RoB domain 5: selection of reported result", "low", ""),
    ("overall_rob", "categorical", "low|some_concerns|high", "Overall risk of bias judgment", "low", ""),
    ("rob_notes", "string", None, "Notes on risk of bias assessment", "Pretest equivalence confirmed", ""),
    # Source Management
    ("human_coder", "string", None, "ID or initials of the human coder", "JD", ""),
    ("coding_date", "date", None, "Date when coding was completed", "2025-06-15", "YYYY-MM-DD format"),
]

EFFECT_SIZES_VARS = [
    # Link
    ("study_id", "integer", None, "Links to STUDY_CHAR.study_id", "1", "Foreign key"),
    ("es_id", "string", None, "Unique effect size identifier", "S001_E01", "Format: S###_E##"),
    ("es_sequence", "integer", None, "Sequence number within study", "1", ""),
    # Sample
    ("n_treatment", "integer", None, "Sample size of treatment group", "60", ""),
    ("n_control", "integer", None, "Sample size of control group", "60", ""),
    # Outcome Classification
    ("outcome_name", "string", None, "Name of the outcome measure", "Final exam score", ""),
    ("outcome_dimension", "categorical", "cognitive|affective|behavioral|metacognitive", "Dimension of the outcome", "cognitive", ""),
    ("blooms_level", "categorical", "remember|understand|apply|analyze|evaluate|create", "Bloom's taxonomy level", "apply", ""),
    ("blooms_order", "categorical", "lower|higher", "Lower-order (remember-apply) vs higher-order (analyze-create)", "lower", ""),
    ("measurement_type", "categorical", "test|questionnaire|rubric|log_data|think_aloud|mixed", "Type of measurement instrument", "test", ""),
    ("measurement_timing", "categorical", "immediate_posttest|delayed_posttest|during_intervention", "When outcome was measured", "immediate_posttest", ""),
    ("reliability_alpha", "numeric", None, "Cronbach's alpha or reliability coefficient", "0.85", "If reported"),
    # Raw Statistics - Post-test
    ("m_treatment_post", "numeric", None, "Treatment group post-test mean", "82.5", ""),
    ("sd_treatment_post", "numeric", None, "Treatment group post-test SD", "12.3", ""),
    ("m_control_post", "numeric", None, "Control group post-test mean", "75.2", ""),
    ("sd_control_post", "numeric", None, "Control group post-test SD", "11.8", ""),
    # Raw Statistics - Pre-test
    ("m_treatment_pre", "numeric", None, "Treatment group pre-test mean (optional)", "70.1", "If available"),
    ("sd_treatment_pre", "numeric", None, "Treatment group pre-test SD (optional)", "10.5", "If available"),
    ("m_control_pre", "numeric", None, "Control group pre-test mean (optional)", "69.8", "If available"),
    ("sd_control_pre", "numeric", None, "Control group pre-test SD (optional)", "10.2", "If available"),
    # Alternative Statistics
    ("reported_stat_type", "categorical", "t|F|chi2|p_only|d|eta_sq|r|OR|other", "Type of statistic reported", "t", "Use when means/SDs unavailable"),
    ("reported_stat_value", "numeric", None, "Value of the reported statistic", "2.45", ""),
    ("reported_df", "numeric", None, "Degrees of freedom", "118", ""),
    ("reported_p", "numeric", None, "Reported p-value", "0.016", ""),
    # Calculated Effect Sizes
    ("hedges_g", "numeric", None, "Hedges' g (bias-corrected SMD)", "0.62", "Primary effect size metric"),
    ("se_g", "numeric", None, "Standard error of Hedges' g", "0.18", ""),
    ("var_g", "numeric", None, "Variance of Hedges' g", "0.032", "se_g squared"),
    ("ci_lower_95", "numeric", None, "Lower bound of 95% CI", "0.27", ""),
    ("ci_upper_95", "numeric", None, "Upper bound of 95% CI", "0.97", ""),
    ("calculation_method", "categorical", "post_means|change_scores|t_to_g|F_to_g|p_to_g|reported_d_to_g|eta_to_g|r_to_g", "Method used to calculate g", "post_means", ""),
    # Source Tracking
    ("source_location", "string", None, "Where in the paper data was found", "Table 3, p.12", ""),
    ("extraction_notes", "string", None, "Notes on data extraction", "Values read from figure using WebPlotDigitizer", ""),
    # Source Management
    ("human_coder", "string", None, "ID or initials of the human coder", "JD", ""),
    ("coding_date", "date", None, "Date of coding", "2025-06-15", "YYYY-MM-DD format"),
    ("verification_status", "categorical", "initial|double_coded|consensus_resolved", "Status of verification", "initial", ""),
]

RAW_EXTRACTION_VARS = [
    # Link
    ("study_id", "integer", None, "Links to STUDY_CHAR.study_id", "1", "Foreign key"),
    ("es_id", "string", None, "Links to EFFECT_SIZES.es_id", "S001_E01", ""),
    ("extraction_id", "string", None, "Unique extraction identifier", "S001_E01_X01", ""),
    # Source Info
    ("page_number", "integer", None, "Page number in the source", "12", ""),
    ("table_or_figure", "string", None, "Table or figure reference", "Table 3", "Free text - table/figure numbers vary"),
    ("section", "categorical", "results|discussion|appendix", "Section of the paper", "results", ""),
    ("exact_quote", "string", None, "Verbatim quote from the paper", "The treatment group scored significantly higher (M=82.5, SD=12.3)", ""),
    # Extracted Values
    ("variable_extracted", "string", None, "Name of the variable being extracted", "m_treatment_post", "Should match a variable name in another sheet"),
    ("value", "string", None, "Extracted value as a string", "82.5", ""),
    ("unit_or_scale", "string", None, "Unit or scale of the value", "0-100 points", ""),
    # Reliability
    ("extraction_confidence", "categorical", "high|medium|low", "Confidence in the extraction", "high", ""),
    ("confidence_notes", "string", None, "Notes on confidence rating", "Clearly reported in table", ""),
]

ES_CALCULATIONS_VARS = [
    # Link
    ("study_id", "integer", None, "Links to STUDY_CHAR.study_id", "1", "Foreign key"),
    ("es_id", "string", None, "Links to EFFECT_SIZES.es_id", "S001_E01", ""),
    # Inputs
    ("input_m1", "numeric", None, "Treatment group mean", "82.5", ""),
    ("input_sd1", "numeric", None, "Treatment group SD", "12.3", ""),
    ("input_n1", "integer", None, "Treatment group n", "60", ""),
    ("input_m2", "numeric", None, "Control group mean", "75.2", ""),
    ("input_sd2", "numeric", None, "Control group SD", "11.8", ""),
    ("input_n2", "integer", None, "Control group n", "60", ""),
    ("input_stat_type", "string", None, "Type of input statistic if not means", "t", ""),
    ("input_stat_value", "numeric", None, "Value of input statistic", "2.45", ""),
    # Calculation
    ("formula_used", "string", None, "Formula or method used for calculation", "d = (M1-M2)/SD_pooled; g = d*J", ""),
    ("cohens_d_raw", "numeric", None, "Raw Cohen's d before correction", "0.61", ""),
    ("correction_factor_J", "numeric", None, "Small-sample correction factor J", "0.9916", "J = 1 - 3/(4*df - 1)"),
    ("hedges_g_final", "numeric", None, "Final Hedges' g value", "0.60", "Should match EFFECT_SIZES.hedges_g"),
    ("se_g_final", "numeric", None, "Final SE of Hedges' g", "0.18", "Should match EFFECT_SIZES.se_g"),
    # Verification
    ("matches_sheet3", "categorical", "yes|no", "Whether values match EFFECT_SIZES sheet", "yes", "Cross-check flag"),
    ("verification_note", "string", None, "Notes on verification", "Values match within rounding", ""),
]

AI_CODING_VARS = [
    # Link
    ("study_id", "integer", None, "Links to STUDY_CHAR.study_id", "1", "Foreign key"),
    ("es_id", "string", None, "Links to EFFECT_SIZES.es_id", "S001_E01", ""),
    ("variable_name", "string", None, "Variable being coded", "genai_tool", ""),
    # Human Reference
    ("human_gold_value", "string", None, "Gold-standard human-coded value", "ChatGPT", ""),
    ("human_gold_source", "string", None, "Source of human gold value", "Table 1, p.5", ""),
    # Claude
    ("claude_value", "string", None, "Value extracted by Claude", "ChatGPT", ""),
    ("claude_confidence", "numeric", None, "Claude's confidence score (0-1)", "0.95", ""),
    ("claude_raw_output", "string", None, "Full raw output from Claude", "Based on the text...", "For audit trail"),
    ("claude_prompt_version", "string", None, "Version of prompt used for Claude", "v2.1", ""),
    # GPT-4o
    ("gpt4o_value", "string", None, "Value extracted by GPT-4o", "ChatGPT", ""),
    ("gpt4o_confidence", "numeric", None, "GPT-4o's confidence score (0-1)", "0.92", ""),
    ("gpt4o_raw_output", "string", None, "Full raw output from GPT-4o", "The paper mentions...", "For audit trail"),
    ("gpt4o_prompt_version", "string", None, "Version of prompt used for GPT-4o", "v2.1", ""),
    # Gemini
    ("gemini_value", "string", None, "Value extracted by Gemini", "ChatGPT", ""),
    ("gemini_confidence", "numeric", None, "Gemini's confidence score (0-1)", "0.88", ""),
    ("gemini_raw_output", "string", None, "Full raw output from Gemini", "According to the study...", "For audit trail"),
    ("gemini_prompt_version", "string", None, "Version of prompt used for Gemini", "v2.1", ""),
    # Consensus
    ("ai_consensus_value", "string", None, "Final consensus value across AI models", "ChatGPT", ""),
    ("ai_consensus_method", "categorical", "majority|mean|weighted", "Method used to reach consensus", "majority", ""),
    ("ai_human_match", "categorical", "exact|within_threshold|mismatch", "Whether AI consensus matches human", "exact", ""),
]

DISCREPANCY_LOG_VARS = [
    # Link
    ("study_id", "integer", None, "Links to STUDY_CHAR.study_id", "1", "Foreign key"),
    ("es_id", "string", None, "Links to EFFECT_SIZES.es_id", "S001_E01", ""),
    ("variable_name", "string", None, "Variable with discrepancy", "genai_tool", ""),
    # Discrepancy
    ("coder1_value", "string", None, "Value from first coder", "ChatGPT", ""),
    ("coder2_value", "string", None, "Value from second coder", "GPT-4", ""),
    ("discrepancy_type", "categorical", "human_human|ai_human|ai_ai", "Type of discrepancy", "human_human", ""),
    ("magnitude", "categorical", "minor|major", "Severity of discrepancy", "minor", ""),
    # Resolution
    ("final_value", "string", None, "Resolved final value", "ChatGPT", ""),
    ("resolution_method", "categorical", "discussion|third_reviewer|source_recheck|calculation_error", "How the discrepancy was resolved", "discussion", ""),
    ("resolved_by", "string", None, "Who resolved the discrepancy", "JD", ""),
    ("resolution_date", "date", None, "Date of resolution", "2025-06-20", "YYYY-MM-DD format"),
]

EXCLUSION_LOG_VARS = [
    ("study_id", "integer", None, "Unique numeric ID of excluded study", "99", ""),
    ("first_author", "string", None, "Last name of the first author", "Jones", ""),
    ("year", "integer", None, "Publication year", "2023", ""),
    ("title", "string", None, "Full title of the paper", "AI in the Classroom", ""),
    ("exclusion_stage", "categorical", "title_abstract|full_text|post_coding", "Stage at which exclusion occurred", "full_text", ""),
    ("exclusion_reason", "categorical", "wrong_population|no_control|non_GenAI|insufficient_data|duplicate|not_peer_reviewed|wrong_design|other", "Reason for exclusion", "no_control", ""),
    ("detailed_rationale", "string", None, "Detailed explanation for exclusion", "No comparison group; single-arm design only", ""),
    ("screener1_decision", "categorical", "include|exclude|uncertain", "First screener's decision", "exclude", ""),
    ("screener2_decision", "categorical", "include|exclude|uncertain", "Second screener's decision", "exclude", ""),
    ("final_decision", "categorical", "include|exclude", "Final inclusion/exclusion decision", "exclude", ""),
]

SEARCH_LOG_VARS = [
    ("database", "string", None, "Database or source searched", "Scopus", ""),
    ("search_date", "date", None, "Date the search was conducted", "2025-05-01", "YYYY-MM-DD format"),
    ("search_string", "string", None, "Exact search query used", '("generative AI" OR "ChatGPT") AND ("higher education")', ""),
    ("results_count", "integer", None, "Number of results returned", "342", ""),
    ("notes", "string", None, "Additional notes on the search", "Filtered to 2020-2025", ""),
]

# Ordered sheet definitions: (sheet_name, variable_list)
SHEETS = [
    ("STUDY_CHAR", STUDY_CHAR_VARS),
    ("EFFECT_SIZES", EFFECT_SIZES_VARS),
    ("RAW_EXTRACTION", RAW_EXTRACTION_VARS),
    ("ES_CALCULATIONS", ES_CALCULATIONS_VARS),
    ("AI_CODING", AI_CODING_VARS),
    ("DISCREPANCY_LOG", DISCREPANCY_LOG_VARS),
    ("EXCLUSION_LOG", EXCLUSION_LOG_VARS),
    ("SEARCH_LOG", SEARCH_LOG_VARS),
]

CODEBOOK_COLUMNS = [
    "variable_name", "sheet", "type", "valid_values",
    "coding_rules", "example", "notes",
]


def apply_header_format(ws, num_cols):
    """Apply formatting to header row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def set_column_widths(ws, headers):
    """Set column widths based on header length, clamped between 12 and 40."""
    for col_idx, header in enumerate(headers, 1):
        width = max(12, min(40, len(header) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_data_validation(ws, col_idx, valid_values_str, max_row=1000):
    """Add dropdown data validation for a column from row 2 to max_row."""
    values = valid_values_str.split("|")
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,  # In openpyxl, False actually shows the dropdown
        showErrorMessage=True,
        errorTitle="Invalid Entry",
        error=f"Value must be one of: {', '.join(values)}",
    )
    col_letter = get_column_letter(col_idx)
    dv.sqref = f"{col_letter}2:{col_letter}{max_row}"
    ws.add_data_validation(dv)


def apply_number_format(ws, col_idx, var_type, max_row=1000):
    """Apply number format for date and numeric columns."""
    col_letter = get_column_letter(col_idx)
    if var_type == "date":
        for row in range(2, max_row + 1):
            ws[f"{col_letter}{row}"].number_format = "YYYY-MM-DD"
    elif var_type == "numeric":
        for row in range(2, max_row + 1):
            ws[f"{col_letter}{row}"].number_format = "0.00##"


def build_data_sheet(wb, sheet_name, variables):
    """Create a data sheet with headers, formatting, validation, and number formats."""
    ws = wb.create_sheet(title=sheet_name)

    # Write headers
    headers = [v[0] for v in variables]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Formatting
    apply_header_format(ws, len(headers))
    set_column_widths(ws, headers)

    # Freeze pane and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Tab color
    if sheet_name in TAB_COLORS:
        ws.sheet_properties.tabColor = TAB_COLORS[sheet_name]

    # Data validation and number formats
    for col_idx, var in enumerate(variables, 1):
        var_name, var_type, valid_values = var[0], var[1], var[2]
        if valid_values is not None and var_type == "categorical":
            add_data_validation(ws, col_idx, valid_values)
        apply_number_format(ws, col_idx, var_type)

    return ws


def build_codebook(wb, all_sheets_vars):
    """Create the CODEBOOK sheet with all variables pre-populated."""
    ws = wb.create_sheet(title="CODEBOOK")

    # Write headers
    for col_idx, header in enumerate(CODEBOOK_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Formatting
    apply_header_format(ws, len(CODEBOOK_COLUMNS))
    set_column_widths(ws, CODEBOOK_COLUMNS)

    # Freeze and filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CODEBOOK_COLUMNS))}1"
    ws.sheet_properties.tabColor = TAB_COLORS["CODEBOOK"]

    # Populate rows
    row = 2
    for sheet_name, variables in all_sheets_vars:
        for var in variables:
            var_name, var_type, valid_values, coding_rule, example, notes = var
            ws.cell(row=row, column=1, value=var_name)
            ws.cell(row=row, column=2, value=sheet_name)
            ws.cell(row=row, column=3, value=var_type)
            ws.cell(row=row, column=4, value=valid_values if valid_values else "")
            ws.cell(row=row, column=5, value=coding_rule)
            ws.cell(row=row, column=6, value=example)
            ws.cell(row=row, column=7, value=notes)

            # Light alternating fill for readability
            if row % 2 == 0:
                light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                for c in range(1, len(CODEBOOK_COLUMNS) + 1):
                    ws.cell(row=row, column=c).fill = light_fill

            row += 1

    return ws


def main():
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Build CODEBOOK first (so it's the first tab)
    build_codebook(wb, SHEETS)

    # 2. Build all data sheets
    for sheet_name, variables in SHEETS:
        build_data_sheet(wb, sheet_name, variables)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Template created successfully: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")

    # Quick self-check
    total_codebook_rows = sum(len(v) for _, v in SHEETS)
    print(f"CODEBOOK rows (excluding header): {total_codebook_rows}")
    for sheet_name, variables in SHEETS:
        print(f"  {sheet_name}: {len(variables)} columns")


if __name__ == "__main__":
    main()
