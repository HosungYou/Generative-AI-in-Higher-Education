#!/usr/bin/env python3
"""
Script to:
1. Add 4 excluded studies to VERIFICATION.xlsx (sheet 5_Excluded_Papers)
2. Regenerate PDF_INDEX.md with correct counts (66 included, 4 excluded)
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Excluded studies to add
EXCLUDED_STUDIES = [
    {
        "Study_ID": 9,
        "Author": "Rommel et al.",
        "Year": 2025,
        "Title": "Development of adaptive and emotionally responsive tutoring",
        "Exclusion_Reason": "No quantitative learning outcome data"
    },
    {
        "Study_ID": 18,
        "Author": "Satoru et al.",
        "Year": 2025,
        "Title": "Generative AI and CEFR Levels Evaluating",
        "Exclusion_Reason": "Assessment tool validation, not intervention study"
    },
    {
        "Study_ID": 39,
        "Author": "Jenny et al.",
        "Year": 2025,
        "Title": "Using Generative AI to Promote Psychological topic",
        "Exclusion_Reason": "Qualitative/descriptive study, no effect sizes"
    },
    {
        "Study_ID": 61,
        "Author": "Pablo et al.",
        "Year": 2024,
        "Title": "Thinking Critically about Scientific Information",
        "Exclusion_Reason": "No control group comparison"
    }
]

# File paths
VERIFICATION_FILE = "data/03_final/GenAI_MetaAnalysis_v5_VERIFICATION.xlsx"
V5_CSV = "data/03_final/GenAI_MetaAnalysis_v5.csv"
PDF_INDEX = "pdfs/PDF_INDEX.md"

def update_verification_xlsx():
    """Add excluded studies to VERIFICATION.xlsx sheet 5_Excluded_Papers"""
    print("📊 Updating VERIFICATION.xlsx...")

    # Load workbook
    wb = load_workbook(VERIFICATION_FILE)

    # Get or create sheet
    if "5_Excluded_Papers" in wb.sheetnames:
        ws = wb["5_Excluded_Papers"]
        print(f"   Found existing sheet '5_Excluded_Papers'")
    else:
        ws = wb.create_sheet("5_Excluded_Papers")
        print(f"   Created new sheet '5_Excluded_Papers'")
        # Add headers
        ws.append(["Study_ID", "Author", "Year", "Title", "Exclusion_Reason"])

    # Get current row count
    start_row = ws.max_row + 1

    # Add excluded studies
    for study in EXCLUDED_STUDIES:
        ws.append([
            study["Study_ID"],
            study["Author"],
            study["Year"],
            study["Title"],
            study["Exclusion_Reason"]
        ])
        print(f"   Added: Study {study['Study_ID']} - {study['Author']} ({study['Year']})")

    # Save workbook
    wb.save(VERIFICATION_FILE)
    print(f"✅ Updated {VERIFICATION_FILE}")
    print(f"   Added {len(EXCLUDED_STUDIES)} excluded studies\n")

def get_included_studies():
    """Extract unique studies from V5 CSV"""
    print("📖 Reading included studies from V5 CSV...")

    df = pd.read_csv(V5_CSV)

    # Get unique studies
    studies = df.groupby('Study_ID').first().reset_index()
    studies = studies[['Study_ID', 'Authors', 'Year', 'Title']].sort_values('Study_ID')

    # Count effect sizes per study
    es_counts = df.groupby('Study_ID').size().reset_index(name='ES_Count')
    studies = studies.merge(es_counts, on='Study_ID')

    print(f"   Found {len(studies)} unique studies")
    print(f"   Total effect sizes: {len(df)}\n")

    return studies

def generate_pdf_index(included_studies):
    """Generate complete PDF_INDEX.md"""
    print("📝 Generating PDF_INDEX.md...")

    total_pdfs = 70  # 66 included + 4 excluded
    included_count = len(included_studies)
    excluded_count = 4
    total_es = included_studies['ES_Count'].sum()

    content = f"""# PDF Index for GenAI Meta-Analysis

Last Updated: {datetime.now().strftime('%Y-%m-%d')}

## Summary
- **Total PDFs**: {total_pdfs}
- **Included in V5**: {included_count} studies ({total_es} effect sizes)
- **Excluded**: {excluded_count} studies

---

## Included Studies ({included_count})

| Study_ID | Author | Year | Title | ES_Count |
|----------|--------|------|-------|----------|
"""

    # Add included studies
    for _, row in included_studies.iterrows():
        author = row['Authors'].split(',')[0] if pd.notna(row['Authors']) else "Unknown"
        year = int(row['Year']) if pd.notna(row['Year']) else "N/A"
        title = row['Title'][:80] + "..." if len(str(row['Title'])) > 80 else row['Title']
        es_count = row['ES_Count']

        content += f"| {row['Study_ID']} | {author} | {year} | {title} | {es_count} |\n"

    # Add excluded studies section
    content += f"""
---

## Excluded Studies ({excluded_count})

| Study_ID | Author | Year | Title | Exclusion_Reason |
|----------|--------|------|-------|------------------|
"""

    for study in EXCLUDED_STUDIES:
        content += f"| {study['Study_ID']} | {study['Author']} | {study['Year']} | {study['Title']} | {study['Exclusion_Reason']} |\n"

    # Add footer
    content += f"""
---

## File Naming Convention

PDFs are named as: `###_FirstAuthor_Year_ShortTitle.pdf`
- Study IDs 9, 18, 39, 61 excluded after manual review
- All included studies have corresponding PDFs in `/pdfs` directory

---

## Notes

1. **V5 Dataset**: Final verified meta-analysis dataset
2. **Exclusions**: Studies excluded due to lack of quantitative outcomes, no control group, or qualitative design
3. **Effect Sizes**: Total {total_es} effect sizes from {included_count} studies
4. **Verification**: All studies manually reviewed for inclusion criteria compliance

---

*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*
*Total: {total_pdfs} PDFs | {included_count} included | {excluded_count} excluded | {total_es} effect sizes*
"""

    # Write to file
    with open(PDF_INDEX, 'w') as f:
        f.write(content)

    print(f"✅ Generated {PDF_INDEX}")
    print(f"   Included: {included_count} studies, {total_es} ES")
    print(f"   Excluded: {excluded_count} studies\n")

def main():
    """Main execution"""
    print("=" * 70)
    print("UPDATING EXCLUDED PAPERS AND REGENERATING PDF_INDEX.md")
    print("=" * 70 + "\n")

    # Step 1: Update VERIFICATION.xlsx
    update_verification_xlsx()

    # Step 2: Get included studies from V5
    included_studies = get_included_studies()

    # Step 3: Generate PDF_INDEX.md
    generate_pdf_index(included_studies)

    # Summary
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"✅ Added 4 excluded studies to VERIFICATION.xlsx")
    print(f"✅ Regenerated PDF_INDEX.md with:")
    print(f"   - {len(included_studies)} included studies")
    print(f"   - {included_studies['ES_Count'].sum()} effect sizes")
    print(f"   - 4 excluded studies")
    print(f"   - Total: 70 PDFs")
    print("=" * 70 + "\n")

if __name__ == "__main__":
    main()
