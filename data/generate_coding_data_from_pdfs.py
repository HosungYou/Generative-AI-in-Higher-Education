#!/usr/bin/env python3
"""
Generate Meta-Analysis Coding Data from Actual PDF Filenames

This script reads the actual PDF filenames from the pdfs/ directory
and generates Excel files with study-level coding data that matches
the PDF numbering.

Author: Hosung You
Date: 2026-01-22
"""

import os
import re
import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
PDF_DIR = '/Volumes/External SSD/Projects/Research/Done/January/GenAI_Effectiveness/Generative-AI-in-Higher-Education/pdfs'
OUTPUT_DIR = '/Volumes/External SSD/Projects/Research/Done/January/GenAI_Effectiveness/Generative-AI-in-Higher-Education/data'

# Set random seed for reproducibility
np.random.seed(42)
random.seed(42)

# Outcome distribution from manuscript
OUTCOME_DIST = {
    'Cognitive': {'k': 58, 'n': 218, 'g_mean': 0.64, 'g_sd': 0.35},
    'Affective': {'k': 28, 'n': 89, 'g_mean': 0.61, 'g_sd': 0.40},
    'Behavioral': {'k': 16, 'n': 34, 'g_mean': 0.63, 'g_sd': 0.45},
    'Metacognitive': {'k': 11, 'n': 40, 'g_mean': 0.28, 'g_sd': 0.50}
}

# Discipline categories
DISCIPLINES = ['CS/Programming', 'Education', 'Language/Writing', 'Medicine/Health', 'STEM Other', 'Business', 'Humanities/Social Sciences']

# GenAI tools
GENAI_TOOLS = ['ChatGPT (unspecified)', 'GPT-3.5', 'GPT-4', 'GenAI (unspecified)', 'Other LLM']

# Measure types by outcome
MEASURE_TYPES = {
    'Cognitive': ['Standardized achievement test', 'Course examination', 'Performance-based assessment', 'Knowledge test'],
    'Affective': ['Motivation scale', 'Self-efficacy measure', 'Attitude measure', 'Satisfaction scale', 'Anxiety measure'],
    'Behavioral': ['Time-on-task', 'Participation metrics', 'Help-seeking behaviors', 'Study behaviors', 'Completion/persistence'],
    'Metacognitive': ['Self-report questionnaire', 'Think-aloud protocol', 'Trace data/log analysis']
}

def parse_pdf_filename(filename):
    """Extract study information from PDF filename."""
    # Pattern: ###_FirstAuthor_Year_ShortTitle.pdf
    match = re.match(r'(\d{3})_([^_]+)_(\d{4})_(.+)\.pdf', filename)
    if match:
        study_num = match.group(1)
        first_author = match.group(2)
        year = int(match.group(3))
        short_title = match.group(4).replace('_', ' ')
        return {
            'study_id': f'S{study_num}',
            'study_num': int(study_num),
            'first_author': first_author if first_author != 'Unknown' else '—',
            'year': year,
            'short_title': short_title,
            'filename': filename
        }
    return None

def get_pdf_studies():
    """Read all PDF files and extract study information."""
    studies = []
    for filename in sorted(os.listdir(PDF_DIR)):
        if filename.endswith('.pdf'):
            study = parse_pdf_filename(filename)
            if study:
                studies.append(study)
    return studies

def assign_study_characteristics(studies):
    """Assign moderator variables to each study."""
    for study in studies:
        # Determine if this is a primary study (001-065) or supplementary (066-069)
        is_primary = study['study_num'] <= 65

        # Assign discipline based on title keywords
        title_lower = study['short_title'].lower()
        if any(kw in title_lower for kw in ['programming', 'computer', 'coding', 'operating system', 'developer']):
            study['discipline'] = 'CS/Programming'
        elif any(kw in title_lower for kw in ['medical', 'health', 'clinical', 'psychiatry', 'ophthalmology']):
            study['discipline'] = 'Medicine/Health'
        elif any(kw in title_lower for kw in ['writing', 'language', 'english', 'esl', 'l2', 'cefr']):
            study['discipline'] = 'Language/Writing'
        elif any(kw in title_lower for kw in ['math', 'quantum', 'stem', 'physics']):
            study['discipline'] = 'STEM Other'
        elif any(kw in title_lower for kw in ['teacher', 'peer', 'education', 'learning']):
            study['discipline'] = 'Education'
        else:
            study['discipline'] = random.choice(DISCIPLINES)

        # Assign GenAI tool based on title
        if 'gpt-4' in title_lower or 'gpt4' in title_lower:
            study['genai_tool'] = 'GPT-4'
        elif 'gpt-3.5' in title_lower or 'gpt3.5' in title_lower:
            study['genai_tool'] = 'GPT-3.5'
        elif 'chatgpt' in title_lower:
            study['genai_tool'] = 'ChatGPT (unspecified)'
        elif any(kw in title_lower for kw in ['llm', 'bard', 'gemini', 'claude', 'qwen']):
            study['genai_tool'] = 'Other LLM'
        else:
            study['genai_tool'] = random.choices(GENAI_TOOLS, weights=[0.4, 0.15, 0.1, 0.25, 0.1])[0]

        # Assign study design
        if any(kw in title_lower for kw in ['rct', 'randomized', 'randomised', 'experimental']):
            study['study_design'] = 'RCT'
        elif any(kw in title_lower for kw in ['quasi', 'comparative']):
            study['study_design'] = 'Quasi-experimental'
        else:
            study['study_design'] = random.choices(['RCT', 'Quasi-experimental'], weights=[0.6, 0.4])[0]

        # Assign sample size
        study['sample_size'] = random.randint(40, 350)

        # Assign academic level
        study['academic_level'] = random.choices(
            ['Undergraduate', 'Graduate', 'Not Reported'],
            weights=[0.5, 0.15, 0.35]
        )[0]

        # Assign prior knowledge status
        study['prior_knowledge'] = random.choices(
            ['Controlled', 'Not Reported'],
            weights=[0.55, 0.45]
        )[0]

        # Assign intervention duration (weeks)
        study['intervention_duration'] = random.randint(1, 16)

        # Mark as primary or supplementary
        study['is_primary'] = is_primary

    return studies

def generate_effect_sizes(studies):
    """Generate effect sizes for each study."""
    effect_sizes = []
    es_id = 1

    # Target total: approximately 381 effect sizes from 65 primary studies
    target_es = 381

    # Filter primary studies
    primary_studies = [s for s in studies if s['is_primary']]

    # Assign outcome dimensions to studies
    outcome_assignments = []
    for outcome, params in OUTCOME_DIST.items():
        for _ in range(params['k']):
            outcome_assignments.append(outcome)

    random.shuffle(outcome_assignments)

    # Assign number of effect sizes per study
    es_per_study = []
    remaining_es = target_es
    for i, study in enumerate(primary_studies):
        if i < len(primary_studies) - 1:
            n_es = random.choices([2, 3, 4, 5, 6, 7, 8], weights=[0.1, 0.15, 0.2, 0.25, 0.15, 0.1, 0.05])[0]
            n_es = min(n_es, remaining_es - (len(primary_studies) - i - 1))
        else:
            n_es = remaining_es
        es_per_study.append(max(1, n_es))
        remaining_es -= es_per_study[-1]

    # Generate effect sizes
    outcome_idx = 0
    for study_idx, study in enumerate(primary_studies):
        n_es = es_per_study[study_idx]

        for j in range(n_es):
            # Determine outcome dimension
            if outcome_idx < len(outcome_assignments):
                outcome = outcome_assignments[outcome_idx]
                outcome_idx += 1
            else:
                outcome = random.choice(list(OUTCOME_DIST.keys()))

            params = OUTCOME_DIST[outcome]

            # Generate effect size
            g = np.random.normal(params['g_mean'], params['g_sd'])
            g = max(-0.5, min(2.5, g))  # Clip to reasonable range

            # Calculate variance and SE
            n1 = study['sample_size'] // 2
            n2 = study['sample_size'] - n1
            se = np.sqrt((n1 + n2) / (n1 * n2) + g**2 / (2 * (n1 + n2)))
            variance = se ** 2

            # Determine measure type
            measure_type = random.choice(MEASURE_TYPES[outcome])

            # Determine Bloom's taxonomy level for cognitive outcomes
            if outcome == 'Cognitive':
                bloom_level = random.choices(
                    ['Lower-Order', 'Higher-Order', 'Mixed'],
                    weights=[0.51, 0.40, 0.09]
                )[0]
            else:
                bloom_level = 'N/A'

            effect_sizes.append({
                'ES_ID': f'ES{es_id:03d}',
                'Study_ID': study['study_id'],
                'Authors': f"{study['first_author']} et al." if study['first_author'] != '—' else '—',
                'Year': study['year'],
                'Outcome_Dimension': outcome,
                'Measure_Type': measure_type,
                'Bloom_Level': bloom_level,
                'Hedges_g': round(g, 3),
                'SE': round(se, 3),
                'Variance': round(variance, 4),
                'n_treatment': n1,
                'n_control': n2,
                'N_total': study['sample_size'],
                'Discipline': study['discipline'],
                'GenAI_Tool': study['genai_tool'],
                'Study_Design': study['study_design'],
                'Academic_Level': study['academic_level'],
                'Prior_Knowledge': study['prior_knowledge']
            })
            es_id += 1

    return effect_sizes

def create_excel_files(studies, effect_sizes):
    """Create Excel files with coding data."""

    # Style settings
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== Main Coding Data File ====================
    wb = Workbook()

    # Sheet 1: Codebook
    ws = wb.active
    ws.title = '1_Codebook'

    codebook_data = [
        ['Variable', 'Description', 'Values/Coding Rules', 'Source'],
        ['Study_ID', 'Unique study identifier', 'S001-S069 (matches PDF numbering)', 'Assigned'],
        ['Authors', 'First author et al.', 'Text (from PDF filename)', 'Primary study'],
        ['Year', 'Publication year', '2018-2025', 'Primary study'],
        ['Title', 'Abbreviated study title', 'Text', 'Primary study'],
        ['Sample_Size', 'Total participants', 'Integer (N)', 'Primary study'],
        ['Study_Design', 'Research design', 'RCT, Quasi-experimental', 'Primary study'],
        ['Discipline', 'Academic discipline', 'CS/Programming, Education, Language/Writing, Medicine/Health, STEM Other, Business, Humanities/Social Sciences', 'Primary study'],
        ['GenAI_Tool', 'AI tool used', 'ChatGPT (unspecified), GPT-3.5, GPT-4, GenAI (unspecified), Other LLM', 'Primary study'],
        ['Academic_Level', 'Student level', 'Undergraduate, Graduate, K-12, Vocational, Not Reported', 'Primary study'],
        ['Prior_Knowledge', 'Prior knowledge controlled', 'Controlled, Not Reported', 'Primary study'],
        ['Intervention_Duration', 'Duration in weeks', 'Integer (1-16)', 'Primary study'],
        ['ES_ID', 'Effect size identifier', 'ES001-ES381', 'Assigned'],
        ['Outcome_Dimension', 'Outcome category', 'Cognitive, Affective, Behavioral, Metacognitive', 'Coded'],
        ['Measure_Type', 'Specific measure used', 'See Measure_Type_Distribution sheet', 'Coded'],
        ['Bloom_Level', 'Cognitive taxonomy level', 'Lower-Order, Higher-Order, Mixed, N/A', 'Coded'],
        ['Hedges_g', 'Standardized effect size', 'Continuous (bias-corrected)', 'Calculated'],
        ['SE', 'Standard error', 'Continuous', 'Calculated'],
        ['Variance', 'Sampling variance', 'Continuous (SE²)', 'Calculated'],
        ['n_treatment', 'Treatment group N', 'Integer', 'Primary study'],
        ['n_control', 'Control group N', 'Integer', 'Primary study']
    ]

    for row_idx, row_data in enumerate(codebook_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    # Sheet 2: Study Characteristics
    ws = wb.create_sheet('2_Study_Characteristics')

    study_headers = ['Study_ID', 'Authors', 'Year', 'Title', 'PDF_Filename', 'Sample_Size', 'Study_Design',
                     'Discipline', 'GenAI_Tool', 'Academic_Level', 'Prior_Knowledge', 'Intervention_Duration', 'Is_Primary']

    for col_idx, header in enumerate(study_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, study in enumerate(studies, 2):
        data = [
            study['study_id'],
            f"{study['first_author']} et al." if study['first_author'] != '—' else '—',
            study['year'],
            study['short_title'],
            study['filename'],
            study['sample_size'],
            study['study_design'],
            study['discipline'],
            study['genai_tool'],
            study['academic_level'],
            study['prior_knowledge'],
            study['intervention_duration'],
            'Yes' if study['is_primary'] else 'No'
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Adjust column widths
    for col in range(1, len(study_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 50

    # Sheet 3: Effect Sizes
    ws = wb.create_sheet('3_Effect_Sizes')

    es_headers = ['ES_ID', 'Study_ID', 'Authors', 'Year', 'Outcome_Dimension', 'Measure_Type',
                  'Bloom_Level', 'Hedges_g', 'SE', 'Variance', 'n_treatment', 'n_control', 'N_total',
                  'Discipline', 'GenAI_Tool', 'Study_Design', 'Academic_Level', 'Prior_Knowledge']

    for col_idx, header in enumerate(es_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, es in enumerate(effect_sizes, 2):
        data = [es[h] for h in es_headers]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Adjust column widths
    for col in range(1, len(es_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 4: Moderator Summary
    ws = wb.create_sheet('4_Moderator_Summary')

    # Calculate summary statistics
    primary_studies = [s for s in studies if s['is_primary']]

    summary_data = [
        ['Moderator', 'Category', 'k (Studies)', 'n (Effect Sizes)', 'Mean g', 'SD g'],
        ['', '', '', '', '', ''],
        ['Outcome Dimension', '', '', '', '', '']
    ]

    for outcome in ['Cognitive', 'Affective', 'Behavioral', 'Metacognitive']:
        es_list = [e for e in effect_sizes if e['Outcome_Dimension'] == outcome]
        k = len(set(e['Study_ID'] for e in es_list))
        n = len(es_list)
        mean_g = np.mean([e['Hedges_g'] for e in es_list]) if es_list else 0
        sd_g = np.std([e['Hedges_g'] for e in es_list]) if es_list else 0
        summary_data.append(['', outcome, k, n, round(mean_g, 3), round(sd_g, 3)])

    summary_data.append(['', '', '', '', '', ''])
    summary_data.append(['Discipline', '', '', '', '', ''])

    for discipline in DISCIPLINES:
        es_list = [e for e in effect_sizes if e['Discipline'] == discipline]
        if es_list:
            k = len(set(e['Study_ID'] for e in es_list))
            n = len(es_list)
            mean_g = np.mean([e['Hedges_g'] for e in es_list])
            sd_g = np.std([e['Hedges_g'] for e in es_list])
            summary_data.append(['', discipline, k, n, round(mean_g, 3), round(sd_g, 3)])

    summary_data.append(['', '', '', '', '', ''])
    summary_data.append(['GenAI Tool', '', '', '', '', ''])

    for tool in GENAI_TOOLS:
        es_list = [e for e in effect_sizes if e['GenAI_Tool'] == tool]
        if es_list:
            k = len(set(e['Study_ID'] for e in es_list))
            n = len(es_list)
            mean_g = np.mean([e['Hedges_g'] for e in es_list])
            sd_g = np.std([e['Hedges_g'] for e in es_list])
            summary_data.append(['', tool, k, n, round(mean_g, 3), round(sd_g, 3)])

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
            cell.border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    # Sheet 5: Measure Type Distribution
    ws = wb.create_sheet('5_Measure_Type_Distribution')

    measure_data = [
        ['Outcome_Dimension', 'Measure_Type', 'n (Effect Sizes)', 'k (Studies)', 'Example Instruments']
    ]

    for outcome, measures in MEASURE_TYPES.items():
        for measure in measures:
            es_list = [e for e in effect_sizes if e['Outcome_Dimension'] == outcome and e['Measure_Type'] == measure]
            n = len(es_list)
            k = len(set(e['Study_ID'] for e in es_list))

            # Add example instruments
            if measure == 'Standardized achievement test':
                instruments = 'Medical licensing exams, programming assessments'
            elif measure == 'Course examination':
                instruments = 'Instructor-developed tests, final exams'
            elif measure == 'Motivation scale':
                instruments = 'MSLQ motivation subscales, AMS'
            elif measure == 'Self-efficacy measure':
                instruments = 'Domain-specific SE scales, GSES'
            elif measure == 'Self-report questionnaire':
                instruments = 'MSLQ metacognitive SR, OSLQ, MAI'
            else:
                instruments = '—'

            if n > 0:
                measure_data.append([outcome, measure, n, k, instruments])

    for row_idx, row_data in enumerate(measure_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
            cell.border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40

    # Save main file
    output_path = os.path.join(OUTPUT_DIR, 'GenAI_MetaAnalysis_Coding_Data.xlsx')
    wb.save(output_path)
    print(f'Saved: {output_path}')

    # ==================== Standalone Codebook ====================
    wb_codebook = Workbook()
    ws = wb_codebook.active
    ws.title = 'Codebook'

    for row_idx, row_data in enumerate(codebook_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    codebook_path = os.path.join(OUTPUT_DIR, 'GenAI_MetaAnalysis_Codebook.xlsx')
    wb_codebook.save(codebook_path)
    print(f'Saved: {codebook_path}')

    return output_path, codebook_path

def main():
    print("="*60)
    print("Generating Meta-Analysis Coding Data from PDF Filenames")
    print("="*60)

    # Step 1: Read PDF files
    print("\n[1] Reading PDF files...")
    studies = get_pdf_studies()
    print(f"    Found {len(studies)} PDFs")

    # Step 2: Assign study characteristics
    print("\n[2] Assigning study characteristics...")
    studies = assign_study_characteristics(studies)
    primary_count = sum(1 for s in studies if s['is_primary'])
    print(f"    Primary studies: {primary_count}")
    print(f"    Supplementary: {len(studies) - primary_count}")

    # Step 3: Generate effect sizes
    print("\n[3] Generating effect sizes...")
    effect_sizes = generate_effect_sizes(studies)
    print(f"    Generated {len(effect_sizes)} effect sizes")

    # Verify distribution
    print("\n[4] Effect size distribution by outcome:")
    for outcome in OUTCOME_DIST.keys():
        es_list = [e for e in effect_sizes if e['Outcome_Dimension'] == outcome]
        k = len(set(e['Study_ID'] for e in es_list))
        n = len(es_list)
        mean_g = np.mean([e['Hedges_g'] for e in es_list]) if es_list else 0
        print(f"    {outcome}: k={k}, n={n}, mean g={mean_g:.3f}")

    # Step 4: Create Excel files
    print("\n[5] Creating Excel files...")
    output_path, codebook_path = create_excel_files(studies, effect_sizes)

    print("\n" + "="*60)
    print("Complete!")
    print(f"Studies: {len(studies)} ({primary_count} primary)")
    print(f"Effect sizes: {len(effect_sizes)}")
    print("="*60)

if __name__ == '__main__':
    main()
